from django.shortcuts import render
from django.http import JsonResponse
from .forms import SuscriptorForm
import openpyxl
from openpyxl.utils import get_column_letter
import os

def index(request):
    return render(request, 'index.html')

def vintage(request):
    return render(request, 'vintage.html')

def minimalismo(request):
    return render(request, 'minimalismo.html')

def sostenible(request):
    return render(request, 'sostenible.html')

def accesorios(request):
    return render(request, 'accesorios.html')

def suscribirse(request):
    if request.method == "POST":
        form = SuscriptorForm(request.POST)
        if form.is_valid():
            form.save()
            
            nombre = form.cleaned_data['nombre']
            correo_electronico = form.cleaned_data['correo_electronico']
            
            file_path = os.path.join(os.path.dirname(__file__), 'suscriptores.xlsx')
            
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
            else:
                workbook = openpyxl.Workbook()
                workbook.active.title = "Suscriptores"
            
            sheet = workbook.active
            
            row = sheet.max_row + 1
            
            sheet[f'A{row}'] = nombre
            sheet[f'B{row}'] = correo_electronico
            
            workbook.save(file_path)
            
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'errors': form.errors})
    return render(request, 'suscribirse.html', {'form': SuscriptorForm()})
